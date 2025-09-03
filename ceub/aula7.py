import random

from prettytable import PrettyTable

from openpyxl import Workbook
from openpyxl.styles import Font, Color, Alignment, PatternFill

contador = 1
while contador <= 5:
    print(contador)
    contador += 1

numero_secreto = random.randint(0, 10)
palpite = 0
tentativas = 0
while palpite != numero_secreto:
    palpite = int(input('Adivinhe o número de 0 a 10:\n'))
    tentativas += 1
    if palpite <  numero_secreto and palpite > -1:
        print('O palpite é menor do que o número')
    elif palpite > numero_secreto and palpite < 11:
        print('O palpite é maior do que o número')
    elif palpite < 0 or palpite > 10:
        print('Digite um número entre 0 e 10')

print(f'Parabéns o número era {numero_secreto} em {tentativas} tentativas')

''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''

tabela = PrettyTable(['Nome', 'Idade', 'Cidade'])
tabela.add_row(["Alice", 25, "São Paulo"])
tabela.add_row(["Bob", 30, "Rio de Janeiro"])
tabela.add_row(["Carol", 22, "Belo Horizonte"])

tabela.align['Nome'] = 'l'
tabela.align['Idade'] = 'c'
tabela.align['Cidade'] = 'r'

tabela.padding_width = 1
tabela.border = True

print(tabela)

''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''

workbook = Workbook()
sheet = workbook.active
sheet.title = "Planilha Formatada"

# Estilo para o cabeçalho
header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF') # Branco
header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid') # Azul
header_alignment = Alignment(horizontal='center', vertical='center')

# Estilo para os dados
data_font = Font(name='Calibri', size=11)
data_alignment = Alignment(horizontal='left', vertical='center')


header_row = ["Nome", "Idade", "Profissão"]
for col_num, column_title in enumerate(header_row, 1):
    cell = sheet.cell(row=1, column=col_num, value=column_title)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment

data = [
    ("Alice", 30, "Engenheira"),
    ("Bob", 25, "Designer"),
    ("Charlie", 35, "Analista de Dados")
]
for row_num, row_data in enumerate(data, 2): # Começamos da linha 2
    for col_num, cell_value in enumerate(row_data, 1):
        cell = sheet.cell(row=row_num, column=col_num, value=cell_value)
        cell.font = data_font
        cell.alignment = data_alignment

workbook.save("planilha_formatada.xlsx")
print("Planilha 'planilha_formatada.xlsx' criada com formatação!")

