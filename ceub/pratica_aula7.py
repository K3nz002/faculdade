# pip install prettytable
from prettytable import PrettyTable
# pip install openpyxl
import openpyxl

tarefas = []

def incluir():
   print('-------Inclusão de Tarefa-------')
   tarefa = input('Digite a tarefa que você quer incluir: ')
   prioridade = int(input('Escolha a prioridade:\n1. Alta\n2. Média\n3. Baixa\n4. Sem prioridade\n'))

   match prioridade:
   
        case 1:
            txt_prio = 'Alta'
        case 2:
            txt_prio = 'Média'
        case 3:
            txt_prio = 'Baixa'
        case 4:
            txt_prio = 'Sem Prioridade'

   tupla = (tarefa, txt_prio)
   tarefas.append(tupla)

def listar():
   
    tabela = PrettyTable(['Número', 'Tarefa', 'Prioridade'])

    numero = 0
    for tarefa, txt_prio in tarefas:
        numero += 1
        tabela.add_row([numero, tarefa, txt_prio])

    tabela.align['Número'] = 'l'
    tabela.align['Tarefa'] = 'r'
    tabela.align['Prioridade'] = 'c'

    tabela.padding_width = 1
    tabela.border = True

    print(tabela)

def excluir():
    listar()
    numero = int(input('Escolha o número da tarefa que você quer excluir:\n'))
    tarefas.pop(numero)

def gerar_excel():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Planilha de Tarefas"

    # Colunas
    sheet.column_dimensions['A'].width = 10
    sheet.column_dimensions['B'].width = 40
    sheet.column_dimensions['C'].width = 15

    # Cabeçalho
    sheet.cell(row = 1, column= 1, value= 'Número')
    sheet.cell(row = 1, column= 2, value= 'Tarefa')
    sheet.cell(row = 1, column= 3, value= 'Prioridade')

    # Listagem de dados
    numero = 0
    for tarefa, txt_prio in tarefas:
        numero += 1
        sheet.cell(row = 2, column= 1, value= numero)
        sheet.cell(row = 2, column= 2, value= tarefa)
        sheet.cell(row = 2, column= 3, value= txt_prio)

    workbook.save("Planilha_Tarefas.xlsx")
    print("Planilha 'Planilha_Tarefas.xlsx' criada com formatação!")

if __name__ == '__main__':
   # Menu
   terminar = False
   while terminar == False:
       print('''
       ------ MENU ------
       1 - Incluir Tarefa
       2 - Excluir Tarefa
       3 - Listar Tarefas
       4 - Gerar Excel
       5 - Sair
       ------------------
       ''')
       opc = input('Qual a opção? ')
       if opc == '1':
           incluir()
       elif opc == '2':
           excluir()
       elif opc == '3':
           listar()
       elif opc == '4':
           gerar_excel()
       elif opc == '5':
           terminar = True
   print('Programa finalizado')